"""
Excel file generation service.
Creates formatted Excel files with multiple sheets based on extracted data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from datetime import datetime

from app.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelGenerator:
    """Generate formatted Excel files from extracted data."""
    
    def __init__(self):
        """Initialize Excel generator with styling configuration."""
        logger.debug("Initializing ExcelGenerator with predefined styles")
        self.wb = None
        
        # Style definitions
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.section_font = Font(bold=True, color="000000")
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center")
        logger.debug("Excel styles configured successfully")
    
    def generate_excel(self, data: Dict[str, Any], output_path: str) -> str:
        """
        Generate complete Excel file with all sheets.
        
        Args:
            data: Extracted and structured data
            output_path: Path to save the Excel file
            
        Returns:
            Path to the generated Excel file
        """
        logger.info("=" * 80)
        logger.info("Starting Excel file generation")
        logger.info(f"Output path: {output_path}")
        
        try:
            self.wb = Workbook()
            # Remove default sheet
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]
            
            logger.info("Creating Excel sheets from extracted data...")
            
            # Generate all sheets with individual logging
            logger.debug("Creating Portfolio Summary sheet...")
            self._create_portfolio_summary_sheet(data.get("portfolio_summary", {}))
            
            logger.debug("Creating Schedule of Investments sheet...")
            self._create_schedule_of_investments_sheet(data.get("schedule_of_investments", []))
            
            logger.debug("Creating Statement of Operations sheet...")
            self._create_statement_of_operations_sheet(data.get("statement_of_operations", []))
            
            logger.debug("Creating Statement of Cashflows sheet...")
            self._create_statement_of_cashflows_sheet(data.get("statement_of_cashflows", {}))
            
            logger.debug("Creating PCAP Statement sheet...")
            self._create_pcap_statement_sheet(data.get("pcap_statement", {}))
            
            logger.debug("Creating Portfolio Company Profile sheet...")
            self._create_portfolio_company_profile_sheet(data.get("portfolio_company_profile", []))
            
            logger.debug("Creating Portfolio Company Financials sheet...")
            self._create_portfolio_company_financials_sheet(data.get("portfolio_company_financials", []))
            
            logger.debug("Creating Footnotes sheet...")
            self._create_footnotes_sheet(data.get("footnotes", []))
            
            logger.debug("Creating Reference Values sheet...")
            self._create_reference_values_sheet(data.get("reference_values", {}))
            
            # Save workbook
            logger.info(f"Saving Excel file with {len(self.wb.sheetnames)} sheets...")
            self.wb.save(output_path)
            
            logger.info(f"Excel file generated successfully | Sheets: {len(self.wb.sheetnames)} | Path: {output_path}")
            logger.debug(f"Sheet names: {', '.join(self.wb.sheetnames)}")
            logger.info("=" * 80)
            
            return output_path
            
        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}", exc_info=True)
            raise Exception(f"Failed to generate Excel file: {str(e)}")
    
    def _create_portfolio_summary_sheet(self, data: Dict[str, Any]):
        """Create Sheet 1: Portfolio Summary."""
        logger.debug("Creating Portfolio Summary sheet...")
        ws = self.wb.create_sheet("Portfolio Summary")
        
        # Headers
        ws['A1'] = "Field"
        ws['B1'] = "Value"
        
        self._apply_header_style(ws, 1, 2)
        
        row = 2
        
        # All fields in order matching expected output
        all_fields = [
            ("General Partner", "general_partner"),
            ("ILPA GP", "ilpa_gp"),
            ("Assets Under Management", "assets_under_management"),
            ("Active Funds", "active_funds"),
            ("Active Portfolio Companies", "active_portfolio_companies"),
            ("Fund Name", "fund_name"),
            ("Fund Currency", "fund_currency"),
            ("Total Commitments", "total_commitments"),
            ("Total Drawdowns", "total_drawdowns"),
            ("Remaining Commitments", "remaining_commitments"),
            ("Net Contributions", "net_contributions"),
            ("NAV", "nav"),
            ("Fair Value", "fair_value"),
            ("Total Number of Investments", "total_investments"),
            ("Realized Investments", "realized_investments"),
            ("Unrealized Investments", "unrealized_investments"),
            ("Total Distributions", "total_distributions"),
            ("- as % of Drawdowns", "distributions_percent_of_drawdowns"),
            ("- as % of Commitments", "distributions_percent_of_commitments"),
            ("DPI", "dpi"),
            ("RVPI", "rvpi"),
            ("TVPI", "tvpi"),
            ("IRR", "irr"),
            ("MOIC", "moic"),
            ("Portfolio Breakdown By Region", None),  # Section header
            ("North America", "north_america_percent"),
            ("Europe", "europe_percent"),
            ("Asia", "asia_percent"),
            ("Other Regions", "other_region_percent"),
            ("Portfolio Breakdown By Industry", None),  # Section header
            ("Consumer Goods", "consumer_goods_percent"),
            ("IT", "it_percent"),
            ("Financials", "financials_percent"),
            ("HealthCare", "healthcare_percent"),
            ("Services", "services_percent"),
            ("Industrials", "industrials_percent"),
            ("Other", "other_industry_percent"),
        ]
        
        filled_fields = 0
        for label, key in all_fields:
            ws[f'A{row}'] = label
            if key is None:
                # Section header - make it bold
                ws[f'A{row}'].font = self.section_font
                ws[f'B{row}'] = ""
            else:
                value = data.get(key, "")
                ws[f'B{row}'] = value
                if value:
                    filled_fields += 1
            row += 1
        
        # Auto-size columns
        self._auto_size_columns(ws)
        logger.debug(f"Portfolio Summary sheet created | Fields: {len(all_fields)} | Filled: {filled_fields}")
    
    def _create_schedule_of_investments_sheet(self, data: List[Dict[str, Any]]):
        """Create Sheet 2: Schedule of Investments."""
        logger.debug("Creating Schedule of Investments sheet...")
        ws = self.wb.create_sheet("Schedule of Investments")
        
        headers = [
            "Company", "Fund", "Reported Date", "Investment Status", "Security Type",
            "Number of Shares", "Fund Ownership %", "Initial Investment Date",
            "Fund Commitment", "Total Invested (A)", "Current Cost (B)", 
            "Reported Value (C)", "Realized Proceeds (D)", "LP Ownership % (Fully Diluted)",
            "Final Exit Date", "Valuation Policy", "Period Change in Valuation",
            "Period Change in Cost", "Unrealized Gains/(Losses)", "Movement Summary",
            "Current Quarter Investment Multiple", "Prior Quarter Investment Multiple",
            "Since Inception IRR"
        ]
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Write data
        for row_idx, investment in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=investment.get("company", ""))
            ws.cell(row=row_idx, column=2, value=investment.get("fund", ""))
            ws.cell(row=row_idx, column=3, value=investment.get("reported_date", ""))
            ws.cell(row=row_idx, column=4, value=investment.get("investment_status", ""))
            ws.cell(row=row_idx, column=5, value=investment.get("security_type", ""))
            ws.cell(row=row_idx, column=6, value=investment.get("number_of_shares", ""))
            ws.cell(row=row_idx, column=7, value=investment.get("fund_ownership_percent", ""))
            ws.cell(row=row_idx, column=8, value=investment.get("initial_investment_date", ""))
            ws.cell(row=row_idx, column=9, value=investment.get("fund_commitment", ""))
            ws.cell(row=row_idx, column=10, value=investment.get("total_invested", ""))
            ws.cell(row=row_idx, column=11, value=investment.get("current_cost", ""))
            ws.cell(row=row_idx, column=12, value=investment.get("reported_value", ""))
            ws.cell(row=row_idx, column=13, value=investment.get("realized_proceeds", ""))
            ws.cell(row=row_idx, column=14, value=investment.get("lp_ownership_percent_fully_diluted", ""))
            ws.cell(row=row_idx, column=15, value=investment.get("final_exit_date", ""))
            ws.cell(row=row_idx, column=16, value=investment.get("valuation_policy", ""))
            ws.cell(row=row_idx, column=17, value=investment.get("period_change_in_valuation", ""))
            ws.cell(row=row_idx, column=18, value=investment.get("period_change_in_cost", ""))
            ws.cell(row=row_idx, column=19, value=investment.get("unrealized_gains_losses", ""))
            ws.cell(row=row_idx, column=20, value=investment.get("movement_summary", ""))
            ws.cell(row=row_idx, column=21, value=investment.get("current_quarter_investment_multiple", ""))
            ws.cell(row=row_idx, column=22, value=investment.get("prior_quarter_investment_multiple", ""))
            ws.cell(row=row_idx, column=23, value=investment.get("since_inception_irr", ""))
        
        self._auto_size_columns(ws)
        logger.debug(f"Schedule of Investments sheet created | Investments: {len(data)} | Columns: {len(headers)}")
    
    def _create_statement_of_operations_sheet(self, data: List[Dict[str, Any]]):
        """Create Sheet 3: Statement of Operations."""
        ws = self.wb.create_sheet("Statement of Operations")
        
        headers = [
            "Period", "Portfolio Interest Income", "Portfolio Dividend Income",
            "Other Interest Earned", "Total Income", "Management Fees, Net",
            "Broken Deal Fees", "Interest", "Professional Fees", "Bank Fees",
            "Advisory Directors' Fees", "Insurance", "Total Expenses",
            "Net Operating Income / (Deficit)", "Net Realized Gain / (Loss) on Investments",
            "Net Change in Unrealized Gain / (Loss) on Investments",
            "Net Realized Gain / (Loss) due to F/X",
            "Net Realized and Unrealized Gain / (Loss) on Investments",
            "Net Increase / (Decrease) in Partners' Capital Resulting from Operations"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        for row_idx, period_data in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=period_data.get("period", ""))
            ws.cell(row=row_idx, column=2, value=period_data.get("portfolio_interest_income", ""))
            ws.cell(row=row_idx, column=3, value=period_data.get("portfolio_dividend_income", ""))
            ws.cell(row=row_idx, column=4, value=period_data.get("other_interest_earned", ""))
            ws.cell(row=row_idx, column=5, value=period_data.get("total_income", ""))
            ws.cell(row=row_idx, column=6, value=period_data.get("management_fees_net", ""))
            ws.cell(row=row_idx, column=7, value=period_data.get("broken_deal_fees", ""))
            ws.cell(row=row_idx, column=8, value=period_data.get("interest", ""))
            ws.cell(row=row_idx, column=9, value=period_data.get("professional_fees", ""))
            ws.cell(row=row_idx, column=10, value=period_data.get("bank_fees", ""))
            ws.cell(row=row_idx, column=11, value=period_data.get("advisory_directors_fees", ""))
            ws.cell(row=row_idx, column=12, value=period_data.get("insurance", ""))
            ws.cell(row=row_idx, column=13, value=period_data.get("total_expenses", ""))
            ws.cell(row=row_idx, column=14, value=period_data.get("net_operating_income_deficit", ""))
            ws.cell(row=row_idx, column=15, value=period_data.get("net_realized_gain_loss_on_investments", ""))
            ws.cell(row=row_idx, column=16, value=period_data.get("net_change_in_unrealized_gain_loss_on_investments", ""))
            ws.cell(row=row_idx, column=17, value=period_data.get("net_realized_gain_loss_due_to_fx", ""))
            ws.cell(row=row_idx, column=18, value=period_data.get("net_realized_and_unrealized_gain_loss_on_investments", ""))
            ws.cell(row=row_idx, column=19, value=period_data.get("net_increase_decrease_in_partners_capital", ""))
        
        self._auto_size_columns(ws)
    
    def _create_statement_of_cashflows_sheet(self, data: Dict[str, Any]):
        """Create Sheet 4: Statement of Cashflows."""
        logger.debug("Creating Statement of Cashflows sheet...")
        ws = self.wb.create_sheet("Statement of Cashflows")
        
        # Define the line items with their field mappings (section, field_name, display_label)
        line_items = [
            ("operating_activities", "section_header", "Cash flows from operating activities"),
            ("operating_activities", "net_increase_decrease_partners_capital", "Net increase/(decrease) in partners' capital"),
            ("operating_activities", "adjustments_to_reconcile", "Adjustments to reconcile net increase/(decrease)"),
            ("operating_activities", "net_realized_gain_loss_investments", "Net realized (gain)/loss on investments"),
            ("operating_activities", "net_change_unrealized_gain_loss", "Net change in unrealized (gain)/loss on investments"),
            ("operating_activities", "changes_in_operating_assets_liabilities", "Changes in operating assets and liabilities"),
            ("operating_activities", "increase_decrease_due_from_affiliates", "(Increase)/decrease in due from affiliates"),
            ("operating_activities", "increase_decrease_due_from_third_party", "(Increase)/decrease in due from third party"),
            ("operating_activities", "increase_decrease_due_from_investment", "(Increase)/decrease in due from investment"),
            ("operating_activities", "purchase_of_investments", "Purchase of investments"),
            ("operating_activities", "proceeds_from_sale_of_investments", "Proceeds from sale of investments"),
            ("operating_activities", "net_cash_provided_by_operating_activities", "Net cash provided by/(used in) operating activities"),
            ("financing_activities", "section_header", "Cash flows from financing activities"),
            ("financing_activities", "capital_contributions", "Capital contributions"),
            ("financing_activities", "distributions", "Distributions"),
            ("financing_activities", "increase_decrease_due_to_limited_partners", "Increase/(decrease) in due to limited partners"),
            ("financing_activities", "increase_decrease_due_to_affiliates", "Increase/(decrease) in due to affiliates"),
            ("financing_activities", "increase_decrease_due_from_limited_partners", "(Increase)/decrease in due from limited partners"),
            ("financing_activities", "proceeds_from_loans", "Proceeds from loans"),
            ("financing_activities", "repayment_of_loans", "Repayment of loans"),
            ("financing_activities", "net_cash_provided_by_financing_activities", "Net cash provided by/(used in) financing activities"),
            ("cash_summary", "net_increase_decrease_cash", "Net increase/(decrease) in cash and cash equivalents"),
            ("cash_summary", "cash_beginning_of_period", "Cash and cash equivalents, beginning of period"),
            ("cash_summary", "cash_end_of_period", "Cash and cash equivalents, end of period"),
            ("supplemental_information", "supplemental_disclosure_header", "Supplemental disclosure of cash flow information"),
            ("supplemental_information", "cash_paid_for_interest", "Cash paid for interest"),
        ]
        
        # Write column headers (Description header + all line items)
        ws.cell(row=1, column=1, value="Description")
        cell = ws.cell(row=1, column=1)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border
        
        for col_idx, (_, _, label) in enumerate(line_items, start=2):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Row labels (periods)
        period_labels = ["Current Period", "Prior Period", "Year to Date"]
        for row_idx, label in enumerate(period_labels, start=2):
            ws.cell(row=row_idx, column=1, value=label)
        
        # Fill data (transposed - rows are periods, columns are line items)
        for col_idx, (section, field, _) in enumerate(line_items, start=2):
            section_data = data.get(section, {})
            field_data = section_data.get(field)
            
            if field_data and isinstance(field_data, dict):
                current = field_data.get("current_period", "")
                prior = field_data.get("prior_period", "")
                ytd = field_data.get("year_to_date", "")
                
                ws.cell(row=2, column=col_idx, value=current if current != 0 else "")  # Current Period
                ws.cell(row=3, column=col_idx, value=prior if prior != 0 else "")      # Prior Period
                ws.cell(row=4, column=col_idx, value=ytd if ytd != 0 else "")          # Year to Date
            else:
                # Section headers or null values
                ws.cell(row=2, column=col_idx, value="")
                ws.cell(row=3, column=col_idx, value="")
                ws.cell(row=4, column=col_idx, value="")
        
        # Freeze panes and format
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "B2"
        ws.column_dimensions['A'].width = 20
        
        self._auto_size_columns(ws)
        logger.debug(f"Statement of Cashflows sheet created | Line items: {len(line_items)}")
    
    def _create_pcap_statement_sheet(self, data: Dict[str, Any]):
        """Create Sheet 5: PCAP Statement."""
        logger.debug("Creating PCAP Statement sheet...")
        ws = self.wb.create_sheet("PCAP Statement")
        
        # Define the PCAP line items with their field mappings (section, field_name, display_label)
        line_items = [
            ("nav_movements", "beginning_nav_net_of_incentive", "Beginning NAV - Net of Incentive Allocation"),
            ("nav_movements", "contributions_cash_non_cash", "Contributions - Cash & Non-Cash"),
            ("nav_movements", "distributions_cash_non_cash", "Distributions - Cash & Non-Cash"),
            ("nav_movements", "total_cash_non_cash_flows", "Total Cash / Non-Cash Flows"),
            ("fees_and_expenses", "management_fees_gross", "(Management Fees - Gross of Offsets, Waivers & Rebates)"),
            ("fees_and_expenses", "management_fee_rebate", "(Management Fee Rebate)"),
            ("fees_and_expenses", "partnership_expenses_total", "(Partnership Expenses - Total)"),
            ("fees_and_expenses", "total_offsets_to_fees_expenses", "Total Offsets to Fees & Expenses"),
            ("fees_and_expenses", "fee_waiver", "Fee Waiver"),
            ("income_and_performance", "interest_income", "Interest Income"),
            ("income_and_performance", "dividend_income", "Dividend Income"),
            ("income_and_performance", "interest_expense", "(Interest Expense)"),
            ("income_and_performance", "other_income_expense", "Other Income/(Expense)"),
            ("income_and_performance", "total_net_operating_income", "Total Net Operating Income / (Expense)"),
            ("income_and_performance", "placement_fees", "(Placement Fees)"),
            ("income_and_performance", "realized_gain_loss", "Realized Gain / (Loss)"),
            ("income_and_performance", "change_in_unrealized_gain_loss", "Change in Unrealized Gain / (Loss)"),
            ("ending_nav_and_commitments", "ending_nav_net_of_incentive", "Ending NAV - Net of Incentive Allocation"),
            ("ending_nav_and_commitments", "incentive_allocation_paid", "Incentive Allocation - Paid During the Period"),
            ("ending_nav_and_commitments", "accrued_incentive_allocation_change", "Accrued Incentive Allocation - Periodic Change"),
            ("ending_nav_and_commitments", "accrued_incentive_allocation_balance", "Accrued Incentive Allocation - Ending Period Balance"),
            ("ending_nav_and_commitments", "ending_nav_gross_of_incentive", "Ending NAV - Gross of Accrued Incentive Allocation"),
            ("ending_nav_and_commitments", "total_commitment", "Total Commitment"),
            ("ending_nav_and_commitments", "beginning_unfunded_commitment", "Beginning Unfunded Commitment"),
            ("ending_nav_and_commitments", "plus_recallable_distributions", "Plus Recallable Distributions"),
            ("ending_nav_and_commitments", "less_expired_released_commitments", "Less Expired/Released Commitments"),
            ("ending_nav_and_commitments", "other_unfunded_adjustment", "+/- Other Unfunded Adjustment"),
            ("ending_nav_and_commitments", "ending_unfunded_commitment", "Ending Unfunded Commitment"),
        ]
        
        # Write column headers (Description header + all line items)
        ws.cell(row=1, column=1, value="Description")
        cell = ws.cell(row=1, column=1)
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.border
        
        for col_idx, (_, _, label) in enumerate(line_items, start=2):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Row labels (periods)
        period_labels = ["Current Period", "Prior Period", "Year to Date"]
        for row_idx, label in enumerate(period_labels, start=2):
            ws.cell(row=row_idx, column=1, value=label)
        
        # Fill data (transposed - rows are periods, columns are line items)
        for col_idx, (section, field, _) in enumerate(line_items, start=2):
            section_data = data.get(section, {})
            field_data = section_data.get(field)
            
            if field_data and isinstance(field_data, dict):
                current = field_data.get("current_period", "")
                prior = field_data.get("prior_period", "")
                ytd = field_data.get("year_to_date", "")
                
                ws.cell(row=2, column=col_idx, value=current if current != 0 else "")  # Current Period
                ws.cell(row=3, column=col_idx, value=prior if prior != 0 else "")      # Prior Period
                ws.cell(row=4, column=col_idx, value=ytd if ytd != 0 else "")          # Year to Date
            else:
                # Null values
                ws.cell(row=2, column=col_idx, value="")
                ws.cell(row=3, column=col_idx, value="")
                ws.cell(row=4, column=col_idx, value="")
        
        # Freeze panes and format
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "B2"
        ws.column_dimensions['A'].width = 20
        
        self._auto_size_columns(ws)
        logger.debug(f"PCAP Statement sheet created | Line items: {len(line_items)}")
    
    def _create_portfolio_company_profile_sheet(self, data: List[Dict[str, Any]]):
        """Create Sheet 6: Portfolio Company Profile."""
        ws = self.wb.create_sheet("Portfolio Company Profile")
        
        headers = [
            "Company Name", "Initial Investment Date", "Industry", "Headquarters",
            "Company Description", "Fund Ownership %", "Investor Group Ownership %",
            "Enterprise Valuation at Closing", "Securities Held", "Ticker Symbol",
            "Investor Group Members", "Management Ownership %", "Board Representation",
            "Board Members", "Investment Commitment", "Invested Capital",
            "Reported Value", "Realized Proceeds", "Investment Multiple",
            "Gross IRR (All Security Types)", "Investment Background",
            "Initial Investment Thesis", "Exit Expectations",
            "Recent Events & Key Initiatives", "Company Assessment",
            "Valuation Methodology", "Risk Assessment / Update"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        for row_idx, company in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=company.get("company_name", ""))
            ws.cell(row=row_idx, column=2, value=company.get("initial_investment_date", ""))
            ws.cell(row=row_idx, column=3, value=company.get("industry", ""))
            ws.cell(row=row_idx, column=4, value=company.get("headquarters", ""))
            ws.cell(row=row_idx, column=5, value=company.get("company_description", ""))
            ws.cell(row=row_idx, column=6, value=company.get("fund_ownership_percent", ""))
            ws.cell(row=row_idx, column=7, value=company.get("investor_group_ownership_percent", ""))
            ws.cell(row=row_idx, column=8, value=company.get("enterprise_valuation_at_closing", ""))
            ws.cell(row=row_idx, column=9, value=company.get("securities_held", ""))
            ws.cell(row=row_idx, column=10, value=company.get("ticker_symbol", ""))
            ws.cell(row=row_idx, column=11, value=company.get("investor_group_members", ""))
            ws.cell(row=row_idx, column=12, value=company.get("management_ownership_percent", ""))
            ws.cell(row=row_idx, column=13, value=company.get("board_representation", ""))
            ws.cell(row=row_idx, column=14, value=company.get("board_members", ""))
            ws.cell(row=row_idx, column=15, value=company.get("investment_commitment", ""))
            ws.cell(row=row_idx, column=16, value=company.get("invested_capital", ""))
            ws.cell(row=row_idx, column=17, value=company.get("reported_value", ""))
            ws.cell(row=row_idx, column=18, value=company.get("realized_proceeds", ""))
            ws.cell(row=row_idx, column=19, value=company.get("investment_multiple", ""))
            ws.cell(row=row_idx, column=20, value=company.get("gross_irr", ""))
            ws.cell(row=row_idx, column=21, value=company.get("investment_background", ""))
            ws.cell(row=row_idx, column=22, value=company.get("initial_investment_thesis", ""))
            ws.cell(row=row_idx, column=23, value=company.get("exit_expectations", ""))
            ws.cell(row=row_idx, column=24, value=company.get("recent_events_key_initiatives", ""))
            ws.cell(row=row_idx, column=25, value=company.get("company_assessment", ""))
            ws.cell(row=row_idx, column=26, value=company.get("valuation_methodology", ""))
            ws.cell(row=row_idx, column=27, value=company.get("risk_assessment_update", ""))
        
        self._auto_size_columns(ws)
    
    def _create_portfolio_company_financials_sheet(self, data: List[Dict[str, Any]]):
        """Create Sheet 7: Portfolio Company Financials."""
        ws = self.wb.create_sheet("Portfolio Company Financials")
        
        headers = [
            "Company", "Company Currency", "Operating Data Date", "Data Type",
            "LTM Revenue", "LTM EBITDA", "Cash", "Book Value", "Gross Debt",
            "1 Year", "2 Years", "3 Years", "4 Years", "5 Years", "After 5 Years",
            "YOY % Growth (Revenue)", "LTM EBITDA (Pro-forma)",
            "YOY % Growth (EBITDA)", "EBITDA Margin", "Total Enterprise Value (TEV)",
            "TEV Multiple", "Total Leverage", "Total Leverage Multiple"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        for row_idx, financials in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=financials.get("company", ""))
            ws.cell(row=row_idx, column=2, value=financials.get("company_currency", ""))
            ws.cell(row=row_idx, column=3, value=financials.get("operating_data_date", ""))
            ws.cell(row=row_idx, column=4, value=financials.get("data_type", ""))
            ws.cell(row=row_idx, column=5, value=financials.get("ltm_revenue", ""))
            ws.cell(row=row_idx, column=6, value=financials.get("ltm_ebitda", ""))
            ws.cell(row=row_idx, column=7, value=financials.get("cash", ""))
            ws.cell(row=row_idx, column=8, value=financials.get("book_value", ""))
            ws.cell(row=row_idx, column=9, value=financials.get("gross_debt", ""))
            ws.cell(row=row_idx, column=10, value=financials.get("debt_1_year", ""))
            ws.cell(row=row_idx, column=11, value=financials.get("debt_2_years", ""))
            ws.cell(row=row_idx, column=12, value=financials.get("debt_3_years", ""))
            ws.cell(row=row_idx, column=13, value=financials.get("debt_4_years", ""))
            ws.cell(row=row_idx, column=14, value=financials.get("debt_5_years", ""))
            ws.cell(row=row_idx, column=15, value=financials.get("debt_after_5_years", ""))
            ws.cell(row=row_idx, column=16, value=financials.get("yoy_percent_growth_revenue", ""))
            ws.cell(row=row_idx, column=17, value=financials.get("ltm_ebitda_pro_forma", ""))
            ws.cell(row=row_idx, column=18, value=financials.get("yoy_percent_growth_ebitda", ""))
            ws.cell(row=row_idx, column=19, value=financials.get("ebitda_margin", ""))
            ws.cell(row=row_idx, column=20, value=financials.get("total_enterprise_value", ""))
            ws.cell(row=row_idx, column=21, value=financials.get("tev_multiple", ""))
            ws.cell(row=row_idx, column=22, value=financials.get("total_leverage", ""))
            ws.cell(row=row_idx, column=23, value=financials.get("total_leverage_multiple", ""))
        
        self._auto_size_columns(ws)
    
    def _create_footnotes_sheet(self, data: List[Dict[str, Any]]):
        """Create Sheet 8: Footnotes."""
        ws = self.wb.create_sheet("Footnotes")
        
        headers = ["Note #", "Note Header", "Operating Data Date", "Description"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        for row_idx, footnote in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=footnote.get("note_number", ""))
            ws.cell(row=row_idx, column=2, value=footnote.get("note_header", ""))
            ws.cell(row=row_idx, column=3, value=footnote.get("operating_data_date", ""))
            ws.cell(row=row_idx, column=4, value=footnote.get("description", ""))
        
        self._auto_size_columns(ws)
    
    def _create_reference_values_sheet(self, data: Dict[str, List[str]]):
        """Create Sheet 9: Reference Values."""
        ws = self.wb.create_sheet("Reference Values")
        
        # Create columns for each reference type
        col = 1
        for key, values in data.items():
            # Write header
            header = key.replace("_", " ").title()
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
            
            # Write values
            if isinstance(values, list):
                for row_idx, value in enumerate(values, start=2):
                    ws.cell(row=row_idx, column=col, value=value)
            
            col += 1
        
        self._auto_size_columns(ws)
    
    def _apply_header_style(self, ws, row: int, num_cols: int):
        """Apply header styling to a row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
    
    def _auto_size_columns(self, ws):
        """Auto-size all columns in a worksheet."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
